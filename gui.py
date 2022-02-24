open_file = open
import enum
import typing
from openpyxl import load_workbook
from unidecode import unidecode
# importing all files  from tkinter
from tkinter import * 
from tkinter import ttk
from enum import Enum
from toolz.itertoolz import partition_all
import requests
import time
# today
from datetime import date

URL_ANAF = 'https://webservicesp.anaf.ro/PlatitorTvaRest/api/v6/ws/tva'
today = date.today().strftime("%Y-%m-%d")
print(today)

class colNames(Enum):
    NUMAR_FACTURA = 'Numar factura'
    DATA = 'Data'
    PERSOANA_JURIDICA = 'Persoana juridica'
    CUI = 'CNP / CUI'
    COD_EXTERN = 'Cod extern'
    PRODUS = 'Comenzi'
    COD_PRODUS = 'Cod produs'
    CANTITATE = 'Cantitate'
    PRET = 'Valoare fara TVA'

class Product_Group(Enum):
    PRODUSE = 'Prouduse'
    PRODUS = 'Comenzi'
    COD_PRODUS = 'Cod produs'
    CANTITATE = 'Cantitate'
    PRET = 'Valoare fara TVA'



  
# import only asksaveasfile from filedialog
# which is used to save file in any extension
from tkinter.filedialog import asksaveasfilename, askopenfilename

root = Tk()
root.geometry('200x200')

# function to call when user press
# the save button, a filedialog will
# open and ask to save file
NoneType = type(None)
def texting(text_change: typing.Union[str, NoneType] = None):
    global text
    text.delete('1.0', END)
    if text_change is not None:
        text.insert('1.0', text_change)

collumn_names = dict()

strFactura = ""
strFirme = ""

def save():
    files = [('All Files', '*.*'),
             ('Text Document', '*.txt')]
    file = asksaveasfilename(filetypes = files, defaultextension = files)
    file_facturi = file + "_FACTURI.txt"
    file_firme = file + "_FIRME.txt"
    with open_file(file_facturi, "w") as f:
        f.write(strFactura)
    
    with open_file(file_firme, "w") as f:
        f.write(strFirme)


    print(file)

def getFromRow(row, col_enum):
    return row[collumn_names[col_enum.value]].value

def date_to_str(date):
    return str(date).split(" ")[0]

lastNumber = 0 # will be modified further
externalCodes = dict()
companiesToQuery = list()

NOTHING = '#N/A'
def getExternalCodeCompany(CUI, externalCode):
    global lastNumber
    if CUI in externalCodes:
        return externalCodes[CUI]
    if externalCode == NOTHING:
        lastNumber += 1
        externalCodes[CUI] = "C" + str(lastNumber).zfill(6)
        #add to companies to query
        companiesToQuery.append(CUI)
    if externalCode != NOTHING:
        externalCodes[CUI] = externalCode
    return externalCodes[CUI]

facturi = []

def iterOnRows(rows):
    second_row = rows[2]
    last_doc = 0
    for row in rows.iter_rows(min_row=2):
        current_doc = getFromRow(row, colNames.NUMAR_FACTURA)
        if last_doc != current_doc:
            last_doc = current_doc
            try:
                factura_curenta
            except NameError:
                pass
            else:
                facturi.append(factura_curenta)
            factura_curenta = dict()
            # facturi.append(factura_curenta)
            factura_curenta[colNames.CUI] = getFromRow(row, colNames.CUI)
            factura_curenta[colNames.COD_EXTERN] = getExternalCodeCompany(factura_curenta[colNames.CUI], getFromRow(row, colNames.COD_EXTERN))
            factura_curenta[colNames.DATA] = date_to_str(getFromRow(row, colNames.DATA))
            factura_curenta[colNames.NUMAR_FACTURA] = getFromRow(row, colNames.NUMAR_FACTURA)
            factura_curenta[Product_Group.PRODUSE] = list()
        produs_curent = dict()
        produs_curent[Product_Group.PRODUS] = getFromRow(row, colNames.PRODUS)
        produs_curent[Product_Group.PRET] = getFromRow(row, colNames.PRET)
        produs_curent[Product_Group.COD_PRODUS] = getFromRow(row, colNames.COD_PRODUS)
        produs_curent[Product_Group.CANTITATE] = getFromRow(row, colNames.CANTITATE)
        factura_curenta[Product_Group.PRODUSE].append(produs_curent)


inputtxt = Text(root, height = 1, width = 20)



def convertFacturiToString(data):
    print(data)
    (anul, luna, day) = date_to_str(data).split("-")
    strFactura = """[InfoPachet]
AnLucru={AnLucru}
LunaLucru={LunaLucru}
Tipdocument=FACTURA IESIRE
TotalFacturi={TotalFacturi}

""".format(
        AnLucru= anul,
        LunaLucru= luna,
        TotalFacturi= len(facturi)
    )
    for index, factura in enumerate(facturi, start=1):
        stornare = "D" if factura[Product_Group.PRODUSE][0][Product_Group.CANTITATE] < 0 else "N"
        strFactura += """[Factura_{iterator}]
NrDoc={NrDoc}
CasaDeMarcat=N
Data={data}
CodClient={CodClient}
TVAINCASARE=N
AutoFactura=N
EmisClient=N
EmisTert=N
Stornare={Stornare}
TipTVA=N
ANULAT=N

TotalArticole={totalArticole}
Operat=D

[Items_{iterator}]
""".format(
            iterator = index,
            NrDoc = factura[colNames.NUMAR_FACTURA],
            data = ".".join( reversed( factura[colNames.DATA].split("-") ) ),
            Stornare = stornare,
            CodClient = factura[colNames.COD_EXTERN],
            totalArticole = len(factura[Product_Group.PRODUSE])
        )
        for index_item, item in enumerate(factura[Product_Group.PRODUSE], start=1):
            item_str = ("{CodProdus};LEI;{Cantitate};{Pret}" if item[Product_Group.COD_PRODUS][0]=='S' else "{CodProdus};BUC;{Cantitate};{Pret};DC").format(
                CodProdus = item[Product_Group.COD_PRODUS],
                Cantitate = item[Product_Group.CANTITATE],
                Pret = item[Product_Group.PRET]
            )
            strFactura += """Item_{iterator}={item}
""".format(
                iterator = index_item,
                item =  item_str
            )
        strFactura += '\n'
    return strFactura

def processRequests():
    firme = ""
    for chunk in partition_all(500, companiesToQuery):
        requestData = [{"cui":company, "data":today} for company in chunk]
        reply = requests.post(URL_ANAF, json= requestData).json()['found']
        for iter in reply:
            ADRESA = unidecode(iter["adresa"])
            DENUMIRE = unidecode(iter["denumire"])
            REG_COMERT = iter.get("nrRegCom", "")
            TVA = iter["scpTVA"]
            CUI = iter["cui"]
            COD_PARTENER = externalCodes[CUI]
            JUDET, LOCALITATE = [" ".join(adress.split()[1:]) for adress in ADRESA.split(", ")[:2]]
            if JUDET == "BUCURESTI":
                LOCALITATE = "SECTOR " + LOCALITATE
            firme += """[ParteneriNoi_{COD}]
Denumire={DENUMIRE}
Localitate={LOCALITATE}
Tara=ROMANIA
SimbolTara=RO
Judet={JUDET}
CodFiscal={CUI}
RegistruComert={REG_COMERT}
Sediu=Sediu
CodExtern={COD}
CodIntern=
TipContabil=Tipic

""".format(
            COD = COD_PARTENER,
            DENUMIRE = DENUMIRE,
            CUI = "RO" + str(CUI) if TVA else CUI,
            REG_COMERT = REG_COMERT,
            LOCALITATE = LOCALITATE,
            JUDET = JUDET
        )
        time.sleep(1)
    return firme

def processing(file):
    global inputtxt
    global lastNumber
    global strFactura
    global strFirme
    texting("Wait...")
    lastCode = inputtxt.get("1.0", END)
    print(lastCode)
    lastNumber = int(lastCode[1:].lstrip("0"))
    print(lastNumber)
    currentFile = load_workbook(file)
    currentSheet = currentFile.active
    first_row = currentSheet[1]
    for index, collumn in enumerate(first_row):
       collumn_names[collumn.value] = index
    iterOnRows(currentSheet)
    data = getFromRow(currentSheet[2], colNames.DATA)
    strFactura = convertFacturiToString(data)
    strFirme = processRequests()
    texting("Save file")

    


def open():
    files = [('All Files', '*.*'), 
             ('excel', '*.xlsx')]
    file = askopenfilename(filetypes = files, defaultextension = files)
    processing(file)
    print(file)


inputtxt.pack(side = TOP, pady = 10)

btnSave = ttk.Button(root, text = 'Save', command = lambda : save())
btnOpen = ttk.Button(root, text = 'Open', command = lambda : open())
btnOpen.pack(side = TOP, pady = 10)
btnSave.pack(side = TOP, pady = 10)
text = Text(root)
text.pack(side = TOP, pady = 15)
texting("open excel")
inputtxt.insert('1.0', "C005186")


mainloop()